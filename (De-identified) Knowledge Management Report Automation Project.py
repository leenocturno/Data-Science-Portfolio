#import libraries and load file
import pandas as pd
from datetime import datetime, timedelta
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from dateutil.relativedelta import relativedelta
excel_file = '2022-01-01 Report.xlsx'
df = pd.read_excel(excel_file)

#filter for documents of interest
df=df.sort_values(by = 'Type', ascending = True)
df.reset_index(drop=True, inplace=True)
df = df[df['Type'].isin(["Form", 
                        "Helping Document", 
                        "Legacy Document", 
                        "SOP", 
                        "Work Instruction"
                        "SLA"]
)]

#add a column that indicates if the periodic review is mandatory.
df.insert(10, 'Periodic Review Determined By', df['Next Periodic Review Date'].apply(
    lambda x: 'Mandatory' if pd.notnull(x) else 'Effective + 5yrs'
))

#import json and add comments
#create your own json file
with open('comments.json','r') as f:
    comment_data = json.load(f)

json_df = pd.DataFrame(comment_data['documents_comments'])
df = pd.merge(df, json_df, left_on = 'Document Number', right_on = 'Number', how = 'left')

#extract date from the comments and add it to the Next Periodic Review Date after adding 2 yeras
date_pattern = r'\d{1,2}/\d{1,2}/\d{4}'

def add_two_years(comment):
    if isinstance(comment, str):
        match = re.search(date_pattern, comment)
        if match:
            date_str = match.group(0)
            date_obj = datetime.strptime(date_str, '%m/%d/%Y')
            new_date_obj = date_obj + relativedelta(years = 2)
            return new_date_obj.strftime('%m/%d/%Y')
        else:
            pass
    else:
        pass
df['Date_from_Comment'] = df['Comment'].apply(add_two_years)
df.loc[df['Date_from_Comment'].notnull(),'Next Periodic Review Date'] = df['Date_from_Comment']
df = df.drop(['Date_from_Comment','Number'], axis = 1)

#Create a copy of the original df
df_copy = df.copy().rename(columns={'Document Number': 'Document #'})
for col in ['Effective Date','Next Periodic Review Date','Obsoleted Date'']:
    df_copy[col] = df_copy[col].dt.date

# Result 1 - Documents with PR due in 6 months for overdue
target_date = datetime.today() + timedelta(days = 100)
condition_pr = (df['Next Periodic Review Date']<=target_date) & (df['Document Status'] != 'Obsoleted')
report_pr = df_copy[condition_pr].drop(['Subtype','Retired Date','Comment'], axis = 1)
report_pr

# Result 2 - Documents undergoing changes
df['Document #'] = df['Version'].astype(int)
report_undergoing_changes = df_copy[df['Version'].apply(lambda x: not x.is_integer())]

#Result 3 - Documents Changed in the last 30 days
thirty_days_ago = datetime.today() - timedelta(days = 30)
condition_changed_30_days_ago = (df['Document Status'].isin(['Effective','Obsoleted'])) & ((df['Effective Date'] > thirty_days_ago) | (df['Obsoleted Date']> thirty_days_ago))
report_changed_30_days_ago = df_copy[condition_changed_30_days_ago]

#Result 4 - Document Owner List
doc_owner = df_copy['Owner'].value_counts().reset_index().rename(columns={'count': 'Documents Owned'})

#Export the results to an excel file
def auto_adjust_columns(filename):
    """
    Auto adjust the columns widths of a given Excel file.
    """
    # Load the workbook
    wb = load_workbook(filename)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)

# Write your dataframes to Excel
with pd.ExcelWriter('Your Report is Here.xlsx') as writer:
    df_copy.to_excel(writer, sheet_name='Documents List', index=False)
    report_pr.to_excel(writer, sheet_name='Periodic Review', index=False)
    report_undergoing_changes.to_excel(writer, sheet_name='Doc Undergoing Changes', index=False)
    report_changed_30_days_ago.to_excel(writer, sheet_name='Updated Documents in 30 days', index=False)
    doc_owner.to_excel(writer, sheet_name= 'Document Owners', index = False)

# Adjust the column widths
auto_adjust_columns('Your Report is Here.xlsx')
